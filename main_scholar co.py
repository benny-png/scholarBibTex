import csv
import openpyxl
import sys
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from scholar_2 import PaperScraper
from tor_proxy import TorProxy
import os

def find_column_index(sheet, column_names):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        return None
    header_row = [col.lower() if col else '' for col in header_row]
    for name in [col.lower() for col in column_names]:
        if name in header_row:
            return header_row.index(name)
    return None

# Load the main Excel file
excel_file = 'All UDSM Units.xlsx'
workbook = openpyxl.load_workbook(excel_file)

# Create a folder to store the college-specific CSV files
output_folder = 'college_data'
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Check if last_notebook_processed.txt exists and read from it
last_notebook_processed = None
if os.path.exists('last_notebook_processed.txt'):
    with open('last_notebook_processed.txt', 'r') as file:
        last_notebook_processed = file.read().strip()

# Get the list of sheet names
sheet_names = workbook.sheetnames[1:]  # Exclude the first sheet

# Find the index to start from
start_index = 0
if last_notebook_processed:
    try:
        start_index = sheet_names.index(last_notebook_processed) + 1
    except ValueError:
        print(f"Warning: Last processed notebook '{last_notebook_processed}' not found. Starting from the beginning.")
        start_index = 0

# Iterate through the sheets in the workbook, starting from the appropriate index
for sheet_name in sheet_names[start_index:]:
    college_name = sheet_name
    print(f"Processing sheet: {college_name}")

    # Find the sheet for the current college
    sheet = workbook[sheet_name]

    # Define the required columns and their possible names
    required_columns = {
        'name': ['full name'],
        'post': ['substantive post'],
        'sex': ['sex'],
        'department': ['department'],
        'status': ['status', 'registered']
    }

    # Find the column indices
    column_indices = {}
    for key, possible_names in required_columns.items():
        index = find_column_index(sheet, possible_names)
        if index is None:
            print(f"Error: Could not find column for {key}. Possible names: {possible_names}")
            continue
        column_indices[key] = index

    print(f"Total rows in the sheet: {sheet.max_row}")
    print("Columns found:")
    for key, index in column_indices.items():
        print(f"{key}: {sheet.cell(row=1, column=index+1).value}")

    print("\nFirst few rows:")
    for row in sheet.iter_rows(min_row=2, max_row=6, values_only=True):
        print(row)

    registered_hyperlinks = []
    statuses_found = set()

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        status = row[column_indices['status']]
        if status:
            statuses_found.add(status)
        else:
            statuses_found.add('n/a')
        if status == 'Registered':
            hyperlink_cell = sheet.cell(row=row_index, column=column_indices['status'] + 1)
            print(f"Row {row_index}: Cell value: {hyperlink_cell.value}, Hyperlink: {hyperlink_cell.hyperlink}")
            if hyperlink_cell.hyperlink:
                hyperlink_address = hyperlink_cell.hyperlink.target + '&view_op=list_works&sortby=pubdate'
                registered_hyperlinks.append([
                    row[column_indices['name']],
                    row[column_indices['post']],
                    row[column_indices['sex']],
                    row[column_indices.get('department', 'N/A')] if column_indices.get('department', None) is not None else 'N/A',
                    status,
                    hyperlink_address
                ])
            else:
                print(f"Row {row_index}: Status is 'Registered' but no hyperlink found")
        elif status:
            print(f"Row {row_index}: Status is '{status}'")
        else:
            print(f"Row {row_index}: Status is not set")

    print(f"\nFound {len(registered_hyperlinks)} registered hyperlinks")
    print(f"Statuses found in the sheet: {', '.join(statuses_found)}")

    if len(registered_hyperlinks) == 0:
        print("No registered hyperlinks found. Moving on to the next college.")
        continue

    # Initialize TorProxy
    tor_proxy = TorProxy()
    proxy_ip = "socks5://localhost:9055"

    # Configure Chrome options for the website
    brave_path = "C:\\Program Files\\BraveSoftware\\Brave-Browser\\Application\\brave.exe"
    options = Options()
    options.binary_location = brave_path
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--start-maximized')
    options.add_argument(f'--proxy-server={proxy_ip}')

    # Open the Chrome WebDriver using TOR proxy
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    def extract_citation_metrics(driver):
        metrics = {'citations': 'N/A', 'h_index': 'N/A', 'i10_index': 'N/A'}
        try:
            table = driver.find_element(By.ID, "gsc_rsb_st")
            rows = table.find_elements(By.XPATH, ".//tbody/tr")
            metrics['citations'] = rows[0].find_elements(By.CLASS_NAME, 'gsc_rsb_std')[0].text
            metrics['h_index'] = rows[1].find_elements(By.CLASS_NAME, 'gsc_rsb_std')[0].text
            metrics['i10_index'] = rows[2].find_elements(By.CLASS_NAME, 'gsc_rsb_std')[0].text
        except Exception as e:
            print(f"Error extracting citation metrics: {e}")
        return metrics

    paper_details = []
    paper_scraper = PaperScraper()

    for hyperlink in registered_hyperlinks:
        print(f"Processing: {hyperlink[0]}")
        driver.get(hyperlink[5])
        citation_metrics = extract_citation_metrics(driver)

        while True:
            try:
                show_more_button = driver.find_element(By.ID, "gsc_bpf_more")
                if show_more_button.is_enabled():
                    driver.execute_script("arguments[0].scrollIntoView(true);", show_more_button)
                    show_more_button.click()
                    time.sleep(2)
                else:
                    break
            except:
                break

        elements = driver.find_elements(By.XPATH, '//a[@class="gsc_a_at"]')
        span_elements = driver.find_elements(By.XPATH, '//span[@class="gsc_a_h gsc_a_hc gs_ibl"]')
        cite_elements = driver.find_elements(By.XPATH, '//a[@class="gsc_a_ac gs_ibl"]')
        
        print(f"Found {len(elements)} papers for {hyperlink[0]}")
        
        # Start the timer
        start_time = time.time()
        for i, element in enumerate(elements):
            year_span = span_elements[i] if i < len(span_elements) else None
            year_of_publication = year_span.text if year_span else "N/A"
            
            cite_span = cite_elements[i] if i < len(cite_elements) else None
            no_of_title_cites = cite_span.text if cite_span else "N/A"

            title = element.text
            link = element.get_attribute('href')

            # Use PaperScraper to get additional details
            details = paper_scraper.scrape_paper_details(link)
            print(details)

            # Renew the Tor connection every 30 seconds
            if time.time() - start_time > 30:
                print("Renewing Tor connection...")
                tor_proxy.renew_connection()
                proxy_ip = tor_proxy.get_ip()
                print(f'previous IP: {proxy_ip}')
                
                # Reset the timer
                start_time = time.time()
            
            paper_detail = {
                'full name': hyperlink[0],
                'substantive post': hyperlink[1],
                'sex': hyperlink[2],
                'department': hyperlink[3],
                'status/registered': hyperlink[4],
                'title': title,
                'year': year_of_publication,
                'link': link,
                'citations': citation_metrics['citations'],
                'h_index': citation_metrics['h_index'],
                'i10_index': citation_metrics['i10_index'],
                'title_cites': no_of_title_cites,
                'authors': details['authors'],
                'journal': details['journal'],
                'volume': details['volume'],
                'pages': details['pages'],
                'booktitle': details['booktitle'],
                'organization': details['organization']
            }

            paper_details.append(paper_detail)

    driver.quit()
    paper_scraper.close()

    # Save the paper details to a CSV file
    csv_file = os.path.join(output_folder, f"research_papers_{college_name}.csv")
    fieldnames = ['full name', 'substantive post', 'sex', 'department', 'status/registered',
                  'title', 'link', 'year', 'citations', 'h_index', 'i10_index', 'title_cites', 
                  'authors', 'journal', 'volume', 'pages', 'booktitle', 'organization']

    with open(csv_file, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(paper_details)

    print(f"Research paper details for {college_name} saved to {csv_file}")

    # Save the last notebook processed to a file
    with open('last_notebook_processed.txt', 'w') as file:
        file.write(college_name)

print("Processing complete.")